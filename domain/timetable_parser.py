import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from typing import List
import re


class TimetableRow:
    def __init__(self, time: str, lessons: str):
        self.__time = time
        self.__lessons = lessons

    @property
    def time(self) -> str:
        return self.__time

    @property
    def lessons(self) -> str:
        return self.__lessons


class WeekdayTimetable:
    def __init__(self, weekday: str):
        self.__weekday = weekday
        self.__timetable = []

    def add_row(self, row: TimetableRow):
        self.__timetable.append(row)

    @property
    def weekday(self) -> str:
        return self.__weekday

    @property
    def timetable(self) -> List[TimetableRow]:
        return self.__timetable


class Timetable:
    def __init__(self):
        self.__timetable = []

    def add_weekday(self, tt: WeekdayTimetable):
        self.__timetable.append(tt)

    def add_row_to_last_weekday(self, row: TimetableRow):
        self.__timetable[-1].add_row(row)

    @property
    def timetable(self) -> List[WeekdayTimetable]:
        return self.__timetable


def get_merged_cell_val(sheet: Worksheet, cell) -> str:
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    value = (
        sheet.cell(rng[0].min_row, rng[0].min_col).value
        if len(rng) != 0
        else cell.value
    )
    if not value:  # value can be None
        value = ""
    value = "\n".join(
        map(
            lambda x: " ".join(x.split()),
            filter(lambda x: x.strip(), value.splitlines()),
        )
    )  # Remove excessive line breaks and spaces
    return value


def get_timetable_for_week_from_worksheet(
    ws: Worksheet, required_col: int
) -> Timetable:
    timetable = Timetable()
    for row in ws.iter_rows(
        3, 44, required_col, required_col
    ):  # Why 42 rows? Because 7 lessons per day, six days a week.
        weekday = ws.cell(row[0].row, 1).value
        if weekday:  # Beginning of a new weekday
            timetable.add_weekday(
                WeekdayTimetable(
                    weekday.replace("\n", "").replace(" ", "").capitalize()
                )
            )  # Weekdays are made to appear vertical with newlines
        time = ws.cell(row[0].row, 2).value
        lessons = get_merged_cell_val(ws, row[0])
        timetable.add_row_to_last_weekday(TimetableRow(time, lessons))
    if len(timetable.timetable) == 6:
        # To account for Sunday
        timetable.add_weekday(WeekdayTimetable("Воскресенье"))
    return timetable


def get_timetable_for_group_from_file(
    filename: str, group: str
) -> Timetable | None:
    workbook = openpyxl.load_workbook(filename)
    for ws in workbook.worksheets:
        for col in ws.iter_cols(1, 100, 2, 2):  # Why 100? No reason.
            if re.match(
                rf"\b{re.escape(group)}\b", col[0].value or "", re.IGNORECASE
            ):
                return get_timetable_for_week_from_worksheet(ws, col[0].column)
    return None
